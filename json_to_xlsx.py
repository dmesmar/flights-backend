#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Módulo para convertir JSON de vuelos a formato Excel (.xlsx)
VERSIÓN SIN columnas de valoración ni nivel de precio
"""

import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def convertir_json_a_xlsx(archivo_json, archivo_xlsx="vuelos.xlsx"):
    try:
        with open(archivo_json, "r", encoding="utf-8") as f:
            vuelos_data = json.load(f)
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo {archivo_json}")
        return None
    except json.JSONDecodeError:
        print(f"❌ Error: El archivo {archivo_json} no es un JSON válido")
        return None

    if not vuelos_data:
        print(f"⚠️ El archivo {archivo_json} está vacío")
        return None

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Vuelos"

        # ── Estilos ─────────────────────────────────────────────
        header_fill  = PatternFill("solid", start_color="1F4E79")
        header_font  = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        cell_font    = Font(size=9, name="Arial")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left",   vertical="center")

        thin   = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        fill_even = PatternFill("solid", start_color="F0F7FF")
        fill_odd  = PatternFill("solid", start_color="FFFFFF")

        # ── Cabeceras (sin valoración) ─────────────────────────
        HEADERS = [
            "Fecha", "Origen", "Destino", "Ruta", "Ranking",
            "Aerolinea", "Salida", "Llegada", "Duracion", "Escalas",
            "Precio (€)",
            "Total vuelos", "Mas barato",
        ]

        COL_WIDTHS = [12, 8, 8, 15, 8, 18, 22, 22, 14, 8, 12, 12, 10]

        for col_idx, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.row_dimensions[1].height = 30

        IDX_PRECIO   = HEADERS.index("Precio (€)") + 1
        IDX_CHEAPEST = HEADERS.index("Mas barato") + 1

        # ── Filas ─────────────────────────────────────────────
        for data_idx, vuelo in enumerate(vuelos_data):
            row_idx = data_idx + 2
            default_fill = fill_even if data_idx % 2 == 0 else fill_odd

            precio_num = float(vuelo["precio"].replace("€", "").strip())

            row_values = [
                vuelo.get("fecha", ""),
                vuelo.get("origen", ""),
                vuelo.get("destino", ""),
                vuelo.get("ruta", ""),
                vuelo.get("ranking", ""),
                vuelo.get("aerolinea", ""),
                vuelo.get("salida", ""),
                vuelo.get("llegada", ""),
                vuelo.get("duracion", ""),
                vuelo.get("escalas", ""),
                precio_num,
                vuelo.get("total_vuelos", ""),
                vuelo.get("mas_barato", ""),
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border

                if col_idx == IDX_PRECIO:
                    cell.number_format = '#,##0 "€"'
                    cell.font      = Font(name="Arial", size=9, bold=True)
                    cell.fill      = default_fill
                    cell.alignment = center_align

                elif col_idx == IDX_CHEAPEST:
                    if value == "✓":
                        cell.font = Font(name="Arial", size=9, bold=True, color="00843D")
                        cell.fill = PatternFill("solid", start_color="E8F5E9")
                    else:
                        cell.font = cell_font
                        cell.fill = default_fill
                    cell.alignment = center_align

                else:
                    cell.font = cell_font
                    cell.fill = default_fill
                    if col_idx in (1, 2, 3, 4, 5, 10, 12):
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align

            ws.row_dimensions[row_idx].height = 16

        # ── Tabla Excel ───────────────────────────────────────
        last_row = len(vuelos_data) + 1
        last_col = get_column_letter(len(HEADERS))
        table = Table(displayName="TablaVuelos", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

        ws.freeze_panes = "A2"

        wb.save(archivo_xlsx)
        print(f"✅ Archivo '{archivo_xlsx}' creado con {len(vuelos_data)} vuelos")
        return archivo_xlsx

    except Exception as e:
        print(f"❌ Error al convertir JSON a XLSX: {str(e)}")
        return None


if __name__ == "__main__":
    convertir_json_a_xlsx("vuelos.json", "vuelos.xlsx")